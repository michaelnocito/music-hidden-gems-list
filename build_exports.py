#!/usr/bin/env python3
"""
build_exports.py — generate the two download files from gems.json.

  * music-hidden-gems.xlsx    formatted Excel workbook (styled header,
                              frozen top row, auto-filter, sized columns)
  * music-hidden-gems-playlist.csv
                              minimal Title/Artist/Album CSV — the format
                              playlist importers (Soundiiz, TuneMyMusic,
                              Playlisty) accept for Apple Music / Spotify

Re-run after gems.json changes:  python build_exports.py
"""

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
gems = json.loads((HERE / "gems.json").read_text(encoding="utf-8"))

# ---------- Excel ----------
wb = Workbook()
ws = wb.active
ws.title = "Music Hidden Gems"

headers = ["Rank", "Song", "Artist", "Genre", "Artist Era", "Length",
           "Album", "Listeners", "Playcount", "Plays per Listener", "Gem Score"]
ws.append(headers)

head_fill = PatternFill("solid", fgColor="0E7490")
head_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = head_fill
    cell.font = head_font
    cell.alignment = Alignment(vertical="center")

for i, g in enumerate(gems, start=1):
    ws.append([i, g["track"], g["artist"], g["genre"], g["decade"],
               g["duration"], g["album"], g["listeners"], g["playcount"],
               g["devotion"], g["gem_score"]])

widths = [6, 42, 26, 18, 10, 8, 34, 12, 12, 10, 10]
for col, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col)].width = w
for row in ws.iter_rows(min_row=2, min_col=8, max_col=9):
    for cell in row:
        cell.number_format = "#,##0"
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:K{len(gems) + 1}"

wb.save(HERE / "music-hidden-gems.xlsx")

# ---------- Playlist CSV ----------
with open(HERE / "music-hidden-gems-playlist.csv", "w", newline="",
          encoding="utf-8-sig") as f:   # BOM so Excel/importers read UTF-8
    w = csv.writer(f)
    w.writerow(["Title", "Artist", "Album"])
    for g in gems:
        w.writerow([g["track"], g["artist"], g["album"]])

print(f"wrote music-hidden-gems.xlsx and music-hidden-gems-playlist.csv "
      f"({len(gems)} songs)")
